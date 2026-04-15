"""
Parse Layer — convert raw files into plain text/markdown.

Routing:
  PDF/image  → Gemini Vision (parse_with_gemini)
  Excel/CSV  → openpyxl   (parse_excel)
"""

import io
import logging
from typing import Optional

import openpyxl

from app.core.ai_clients import gemini_parse_pdf, gemini_parse_image

logger = logging.getLogger(__name__)

PDF_EXTRACT_PROMPT = """Extract ALL content from this document with maximum fidelity.

Instructions:
1. Preserve all tables in markdown format (use | separators)
2. Keep all numbers, dates, percentages exactly as written
3. Preserve Vietnamese text accurately
4. For financial statements: keep every line item and its value
5. Return raw extracted content only — no summary, no commentary

Output the full extracted text below:"""

IMAGE_EXTRACT_PROMPT = """Extract all text visible in this image.

Instructions:
1. Preserve all numbers and figures exactly
2. Recreate tables in markdown format
3. Keep Vietnamese text accurate
4. Include all visible text including headers, footers, labels

Output the extracted text:"""


async def parse_with_gemini(file_bytes: bytes, mime_type: str) -> str:
    """
    Parse a document file using Gemini Vision.
    Supports: PDF, PNG, JPEG, WEBP.
    Returns extracted markdown text.
    """
    logger.info(f"[PARSER] parse_with_gemini — mime={mime_type}, size={len(file_bytes)//1024}KB")

    if mime_type == "application/pdf":
        text, tokens = await gemini_parse_pdf(file_bytes, PDF_EXTRACT_PROMPT)
    elif mime_type in ("image/png", "image/jpeg", "image/jpg", "image/webp"):
        text, tokens = await gemini_parse_image(file_bytes, mime_type, IMAGE_EXTRACT_PROMPT)
    else:
        raise ValueError(f"Unsupported mime type for Gemini parser: {mime_type}")

    logger.info(f"[PARSER] extracted {len(text)} chars, tokens={tokens}")
    return text


def parse_excel(file_bytes: bytes) -> str:
    """
    Parse Excel (.xlsx/.xls) to plain text using openpyxl.
    Each sheet becomes a markdown section with a table.
    """
    logger.info(f"[PARSER] parse_excel — size={len(file_bytes)//1024}KB")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        lines = [f"\n## Sheet: {sheet_name}\n"]
        # Build markdown table
        for i, row in enumerate(rows):
            cells = [str(c) if c is not None else "" for c in row]
            line = "| " + " | ".join(cells) + " |"
            lines.append(line)
            if i == 0:
                # Header separator
                lines.append("| " + " | ".join(["---"] * len(cells)) + " |")

        sections.append("\n".join(lines))

    result = "\n\n".join(sections)
    logger.info(f"[PARSER] parse_excel — {len(wb.sheetnames)} sheets, {len(result)} chars")
    return result


def detect_mime_type(filename: str, content_type: Optional[str] = None) -> str:
    """Detect MIME type from filename extension or provided content_type."""
    if content_type and content_type not in ("application/octet-stream", ""):
        return content_type

    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    mime_map = {
        "pdf": "application/pdf",
        "png": "image/png",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "webp": "image/webp",
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xls": "application/vnd.ms-excel",
        "csv": "text/csv",
    }
    return mime_map.get(ext, "application/octet-stream")
