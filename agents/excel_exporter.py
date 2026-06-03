"""Excel exporter — openpyxl model for pro package."""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


def _header_fill(primary: str = "#1e3a8a") -> PatternFill:
    hex_color = primary.lstrip("#")
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> Border:
    thin = Side(style="thin", color="D1D5DB")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_section_title(ws, row: int, title: str, cols: int = 8,
                          primary: str = "1E3A8A"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=primary.lstrip("#"))
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _write_row(ws, row: int, data: list, bold_first: bool = False,
               fill_color: str = None, num_fmt: str = None):
    for j, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=j, value=val)
        cell.alignment = Alignment(horizontal="right" if j > 1 else "left", vertical="center")
        cell.border = _thin_border()
        if bold_first and j == 1:
            cell.font = Font(bold=True, size=10)
        else:
            cell.font = Font(size=10)
        if fill_color:
            cell.fill = PatternFill("solid", fgColor=fill_color)
        if num_fmt and j > 1:
            cell.number_format = num_fmt


def export_excel(payload: dict, out_path: str,
                 config: dict = None, brand: dict = None) -> str:
    config = config or {}
    brand = brand or {}
    colors = config.get("style", {}).get("colors", {})
    primary = (brand.get("primary") or colors.get("primary", "#1e3a8a")).lstrip("#")
    accent = colors.get("positive", "#10b981").lstrip("#")
    excel_font = config.get("style", {}).get("font", {}).get("excel_family", "Calibri")

    financials = payload.get("financials", {})
    company = financials.get("company", {})
    income_cur = financials.get("income_statement", {}).get("current", {})
    income_prev = financials.get("income_statement", {}).get("previous", {})
    bs_cur = (financials.get("balance_sheet", {}).get("current") or {})
    cf_cur = financials.get("cash_flow", {}).get("current", {})
    unit = financials.get("unit", "triệu đồng")
    period_cur = financials.get("period", {}).get("current", {}).get("label", "Hiện tại")
    period_prev = financials.get("period", {}).get("previous", {}).get("label", "Trước")

    proj_data = payload.get("projection", {})
    proj = proj_data.get("projection", proj_data) if isinstance(proj_data, dict) else {}
    projections = proj.get("projections", [])

    val = payload.get("valuation", {})
    vd = val.get("valuation", val) if "valuation" in val else val
    dcf = vd.get("dcf", {})
    multiples = vd.get("multiples", {})
    sensitivity = vd.get("sensitivity", {})
    assumptions = vd.get("assumptions", {})
    summary = vd.get("summary", {})
    football_field = vd.get("football_field", [])

    ratios = payload.get("ratios", {})
    r = ratios.get("ratios", {})

    wb = Workbook()

    # ── Sheet 1: Overview ─────────────────────────────────────────────────────
    ws_ov = wb.active
    ws_ov.title = "Overview"
    ws_ov.column_dimensions["A"].width = 30
    ws_ov.column_dimensions["B"].width = 22
    ws_ov.column_dimensions["C"].width = 22
    ws_ov.row_dimensions[1].height = 25

    _write_section_title(ws_ov, 1, f"VALUATION OVERVIEW — {company.get('name', '')}", 3, primary)
    headers = ["Chỉ tiêu", period_cur, period_prev]
    for j, h in enumerate(headers, 1):
        cell = ws_ov.cell(row=2, column=j, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor="E5E7EB")
        cell.border = _thin_border()

    income_rows = [
        ("Doanh thu thuần", income_cur.get("net_revenue"), income_prev.get("net_revenue")),
        ("COGS", income_cur.get("cogs"), income_prev.get("cogs")),
        ("Lợi nhuận gộp", income_cur.get("gross_profit"), income_prev.get("gross_profit")),
        ("Chi phí bán hàng", income_cur.get("selling_expense"), income_prev.get("selling_expense")),
        ("Chi phí QLDN", income_cur.get("admin_expense"), income_prev.get("admin_expense")),
        ("EBIT", income_cur.get("operating_profit"), income_prev.get("operating_profit")),
        ("Chi phí lãi vay", income_cur.get("interest_expense"), income_prev.get("interest_expense")),
        ("Lợi nhuận ròng", income_cur.get("net_profit_after_tax"), income_prev.get("net_profit_after_tax")),
    ]
    cur_a = bs_cur.get("assets", {}) or {}
    cur_l = bs_cur.get("liabilities", {}) or {}
    cur_e = bs_cur.get("equity", {}) or {}

    for i, (label, cur, prev) in enumerate(income_rows, 3):
        bg = "F8FAFC" if i % 2 == 0 else "FFFFFF"
        _write_row(ws_ov, i, [label, cur, prev], bold_first=True,
                   fill_color=bg, num_fmt="#,##0")

    row = len(income_rows) + 3
    ws_ov.cell(row=row, column=1, value=f"Đơn vị: {unit}").font = Font(italic=True, size=9, color="6B7280")

    # ── Sheet 2: Projections ──────────────────────────────────────────────────
    ws_proj = wb.create_sheet("Projections")
    ws_proj.column_dimensions["A"].width = 28
    for i in range(len(projections)):
        ws_proj.column_dimensions[get_column_letter(i + 2)].width = 16

    _write_section_title(ws_proj, 1, "DỰ PHÓNG TÀI CHÍNH 5 NĂM", len(projections) + 1, primary)
    year_labels = ["Chỉ tiêu"] + [p.get("year_label", f"Y{p.get('year_index',i+1)}") for i, p in enumerate(projections)]
    for j, h in enumerate(year_labels, 1):
        cell = ws_proj.cell(row=2, column=j, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor="E5E7EB")
        cell.border = _thin_border()

    proj_metrics = [
        ("Doanh thu", "revenue"),
        ("Tăng trưởng %", "growth_pct"),
        ("COGS", "cogs"),
        ("Lợi nhuận gộp", "gross_profit"),
        ("EBITDA", "ebitda"),
        ("EBITDA Margin %", "ebitda_margin_pct"),
        ("EBIT", "ebit"),
        ("Lợi nhuận ròng", "net_income"),
        ("Net Margin %", "net_margin_pct"),
        ("CAPEX", "capex"),
        ("ΔWC", "change_in_wc"),
        ("FCFF", "fcff"),
    ]
    for i, (label, key) in enumerate(proj_metrics, 3):
        bg = "F8FAFC" if i % 2 == 0 else "FFFFFF"
        row_data = [label] + [p.get(key) for p in projections]
        fmt = "0.0%" if "pct" in key or key == "growth_pct" else "#,##0"
        if "pct" in key or key == "growth_pct":
            row_data = [label] + [(p.get(key) or 0) / 100 for p in projections]
        _write_row(ws_proj, i, row_data, bold_first=True, fill_color=bg, num_fmt=fmt)

    # ── Sheet 3: Valuation ────────────────────────────────────────────────────
    ws_val = wb.create_sheet("Valuation")
    ws_val.column_dimensions["A"].width = 35
    ws_val.column_dimensions["B"].width = 22
    ws_val.column_dimensions["C"].width = 22

    _write_section_title(ws_val, 1, "ĐỊNH GIÁ", 3, primary)

    def _val_row(ws, row, label, value, note=""):
        ws.cell(row=row, column=1, value=label).font = Font(size=10)
        cell_v = ws.cell(row=row, column=2, value=value)
        cell_v.font = Font(size=10, bold=True)
        cell_v.number_format = "#,##0"
        ws.cell(row=row, column=3, value=note).font = Font(size=9, color="6B7280", italic=True)
        for j in range(1, 4):
            ws.cell(row=row, column=j).border = _thin_border()
            bg = "F8FAFC" if row % 2 == 0 else "FFFFFF"
            ws.cell(row=row, column=j).fill = PatternFill("solid", fgColor=bg)

    r_idx = 2
    _write_section_title(ws_val, r_idx, "ASSUMPTIONS", 3, primary)
    r_idx += 1
    _val_row(ws_val, r_idx, "WACC (%)", assumptions.get("wacc_pct", 0))
    r_idx += 1
    _val_row(ws_val, r_idx, "Terminal Growth (%)", assumptions.get("terminal_growth_pct", 0))
    r_idx += 1
    _val_row(ws_val, r_idx, "EV/EBITDA Multiple", assumptions.get("ev_ebitda_multiple", 0))
    r_idx += 1
    _val_row(ws_val, r_idx, "P/E Multiple", assumptions.get("pe_multiple", 0))
    r_idx += 1
    _val_row(ws_val, r_idx, "P/B Multiple", assumptions.get("pb_multiple", 0))
    r_idx += 1
    _val_row(ws_val, r_idx, "Minority Discount (%)", assumptions.get("minority_discount_pct", 0))
    r_idx += 2

    _write_section_title(ws_val, r_idx, "DCF VALUATION", 3, primary)
    r_idx += 1
    _val_row(ws_val, r_idx, "PV FCFF", dcf.get("pv_fcff", 0), unit)
    r_idx += 1
    _val_row(ws_val, r_idx, "Terminal Value PV", dcf.get("pv_terminal_value", 0), unit)
    r_idx += 1
    _val_row(ws_val, r_idx, "Enterprise Value", dcf.get("enterprise_value", 0), unit)
    r_idx += 1
    _val_row(ws_val, r_idx, "Net Debt", dcf.get("net_debt", 0), unit)
    r_idx += 1
    ws_val.cell(row=r_idx, column=1, value="Equity Value").font = Font(size=11, bold=True)
    ws_val.cell(row=r_idx, column=2, value=dcf.get("equity_value", 0)).font = Font(size=11, bold=True, color=primary)
    ws_val.cell(row=r_idx, column=2).number_format = "#,##0"
    ws_val.cell(row=r_idx, column=3, value=unit).font = Font(size=9, color="6B7280")
    r_idx += 2

    _write_section_title(ws_val, r_idx, "FOOTBALL FIELD", 3, primary)
    r_idx += 1
    for row in football_field:
        _write_row(ws_val, r_idx,
                   [row.get("method", ""), row.get("low", 0), row.get("mid", 0), row.get("high", 0)],
                   bold_first=True, num_fmt="#,##0")
        r_idx += 1
    r_idx += 1

    _write_section_title(ws_val, r_idx, "SUMMARY", 3, primary)
    r_idx += 1
    _val_row(ws_val, r_idx, "Fair Value Low", summary.get("fair_value_low", 0), unit)
    r_idx += 1
    _val_row(ws_val, r_idx, "Fair Value Mid", summary.get("fair_value_mid", 0), unit)
    r_idx += 1
    _val_row(ws_val, r_idx, "Fair Value High", summary.get("fair_value_high", 0), unit)

    # ── Sheet 4: Sensitivity ─────────────────────────────────────────────────
    ws_sens = wb.create_sheet("Sensitivity")
    _write_section_title(ws_sens, 1, "SENSITIVITY MATRIX (WACC × g → Equity Value)", 8, primary)

    wacc_range = sensitivity.get("wacc_range", [])
    g_range = sensitivity.get("g_range", [])
    matrix = sensitivity.get("matrix", [])

    ws_sens.cell(row=2, column=1, value="WACC \\ g →").font = Font(bold=True, size=9)
    for j, g in enumerate(g_range, 2):
        cell = ws_sens.cell(row=2, column=j, value=f"{g:.1f}%")
        cell.font = Font(bold=True, size=9)
        cell.fill = PatternFill("solid", fgColor="E5E7EB")
        cell.border = _thin_border()
        ws_sens.column_dimensions[get_column_letter(j)].width = 14

    for i, (wacc, row) in enumerate(zip(wacc_range, matrix), 3):
        ws_sens.cell(row=i, column=1, value=f"{wacc:.1f}%").font = Font(bold=True, size=9)
        for j, val_n in enumerate(row, 2):
            cell = ws_sens.cell(row=i, column=j, value=val_n)
            cell.number_format = "#,##0"
            cell.font = Font(size=9)
            cell.border = _thin_border()
            if matrix:
                flat = [v for r_ in matrix for v in r_ if v]
                if flat:
                    vmin, vmax = min(flat), max(flat)
                    intensity = (val_n - vmin) / (vmax - vmin) if vmax != vmin else 0.5
                    r_int = int(255 - intensity * 100)
                    g_int = int(150 + intensity * 80)
                    b_int = int(150 - intensity * 50)
                    hex_bg = f"{max(0,min(255,r_int)):02X}{max(0,min(255,g_int)):02X}{max(0,min(255,b_int)):02X}"
                    cell.fill = PatternFill("solid", fgColor=hex_bg)

    # ── Sheet 5: Ratios ───────────────────────────────────────────────────────
    ws_rat = wb.create_sheet("Ratios")
    ws_rat.column_dimensions["A"].width = 28
    ws_rat.column_dimensions["B"].width = 16
    ws_rat.column_dimensions["C"].width = 14

    _write_section_title(ws_rat, 1, "TỶ SỐ TÀI CHÍNH", 3, primary)
    r_idx = 2
    for group_name, group_key in [("Thanh khoản", "liquidity"),
                                    ("Đòn bẩy", "leverage"),
                                    ("Lợi nhuận", "profitability"),
                                    ("Hiệu quả", "efficiency")]:
        _write_section_title(ws_rat, r_idx, group_name, 3, "6B7280")
        r_idx += 1
        group = r.get(group_key, {})
        for k, rd in group.items():
            val = rd.get("value")
            rating = rd.get("rating", "n/a")
            ws_rat.cell(row=r_idx, column=1, value=k).font = Font(size=10)
            ws_rat.cell(row=r_idx, column=2, value=val).font = Font(size=10, bold=True)
            ws_rat.cell(row=r_idx, column=2).number_format = "0.00"
            ws_rat.cell(row=r_idx, column=3, value=rating).font = Font(
                size=10, bold=True,
                color={"good": "10B981", "warning": "F59E0B", "poor": "EF4444"}.get(rating, "6B7280")
            )
            for j in range(1, 4):
                ws_rat.cell(row=r_idx, column=j).border = _thin_border()
            r_idx += 1
        r_idx += 1

    wb.save(out_path)
    return out_path
